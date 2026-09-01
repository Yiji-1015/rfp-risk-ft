#!/usr/bin/env python3
"""단어와 주 라벨의 상관을 세어 마스킹 후보 검토용 표를 만든다.

모델이 무엇을 보고 있는지(`model_explanation_phrases.csv`)와 데이터에 실제로 무엇이
있는지는 다른 질문이다. 이 스크립트는 후자를 센다. 모델을 학습하지 않고 `model_text`의
단어 등장 여부만으로 라벨 분포를 집계하므로, 설명 보고서의 기여도와 나란히 놓고
"모델이 과하게 본 단어"와 "데이터에 원래 있는 신호"를 가를 수 있다.

평가 대상은 동결 앵커 100건을 제외한 924건이다(결정 25).
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Sequence

from scipy.stats import chi2_contingency

from scripts.evaluation.folds import evaluation_excluded_uids
from scripts.labeling.label_dataset import (
    DATASET_VERSION_ENV,
    DEFAULT_DATASET_KEY,
    get_model_text,
    load_label_dataset,
)

ROOT = Path(__file__).resolve().parents[2]
LABELS = ("통상수용", "견적반영", "계약·질의검토")

# 명사 뒤 조사만 뗀다. 형태소 분석기를 붙이지 않으므로 근사이며, 남는 어간이 두 글자
# 미만이면 떼지 않는다("정의"의 "의"를 조사로 오인하지 않기 위해서다).
# ponytail: 정규식 근사. 어미 변화까지 묶어야 하면 형태소 분석기로 올린다.
_JOSA = (
    "에서의", "으로써", "으로서", "에게서", "이라도", "으로", "에서", "에게",
    "까지", "부터", "보다", "라도", "이나", "와의", "과의", "은", "는", "이",
    "가", "을", "를", "의", "에", "도", "만", "나", "와", "과", "로",
)


def stem_approx(token: str) -> str:
    """조사를 뗀 어간 근사. 엑셀에서 조사 변이를 묶어 보기 위한 보조 열이다."""
    for josa in _JOSA:
        if token.endswith(josa) and len(token) - len(josa) >= 2:
            return token[: -len(josa)]
    return token


def tokenize(text: str) -> set[str]:
    """설명 보고서와 같은 규칙으로 단어를 자른다.

    `explanation_viewer._containing_phrases`가 쓰는 정의를 그대로 따른다. 공백으로
    가르고 양끝 기호를 떼며 2~20자만 남긴다. 같은 표에서 두 값을 비교하려면 단어
    경계가 같아야 한다.
    """
    tokens = set()
    for match in re.finditer(r"\S+", text):
        token = re.sub(r"^[^0-9A-Za-z가-힣]+|[^0-9A-Za-z가-힣]+$", "", match.group())
        if 2 <= len(token) <= 20:
            tokens.add(token)
    return tokens


def count_associations(rows: Sequence[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """단어별 등장 건수·문서 수와 라벨 분포를 센다."""
    counts: dict[str, Counter] = defaultdict(Counter)
    documents: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        label = row["primary_action"]
        for token in tokenize(get_model_text(row)):
            counts[token][label] += 1
            documents[token].add(row["document_id"])
    return {
        token: {"counts": counter, "documents": documents[token]}
        for token, counter in counts.items()
    }


def load_model_contributions(path: Path) -> dict[str, dict[str, float]]:
    """설명 보고서의 라벨별 기여도를 단어 단위로 읽는다. 없으면 빈 표를 준다."""
    if not path.exists():
        return {}
    contributions: dict[str, dict[str, float]] = defaultdict(dict)
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for record in csv.DictReader(handle):
            contributions[record["phrase"]][record["label"]] = float(
                record["total_contribution"]
            )
    return contributions


def build_table(
    rows: Sequence[dict[str, Any]],
    contributions: dict[str, dict[str, float]],
    *,
    min_count: int,
) -> list[dict[str, Any]]:
    """단어 한 줄짜리 검토표를 만든다. 정렬은 모델 기여 합 내림차순이다."""
    total = len(rows)
    base = Counter(row["primary_action"] for row in rows)
    table = []
    for token, stats in count_associations(rows).items():
        counter, documents = stats["counts"], stats["documents"]
        present = sum(counter.values())
        if present < min_count:
            continue

        record: dict[str, Any] = {
            "단어": token,
            "어간근사": stem_approx(token),
            "총건수": present,
            "문서수": len(documents),
        }
        lifts = {}
        for label in LABELS:
            share = counter[label] / present
            expected = base[label] / total
            lifts[label] = share / expected if expected else 0.0
            record[f"{label}_건수"] = counter[label]
            record[f"{label}_비율"] = round(share, 4)
            record[f"{label}_lift"] = round(lifts[label], 3)

        top = max(LABELS, key=lambda label: lifts[label])
        record["최대lift_라벨"] = top
        record["최대lift"] = round(lifts[top], 3)
        record["chi2_p"] = _chi2_p(counter, base, present, total)

        contribution = contributions.get(token, {})
        for label in LABELS:
            record[f"모델기여_{label}"] = round(contribution.get(label, 0.0), 4)
        record["모델기여_합"] = round(sum(contribution.values()), 4)
        table.append(record)

    table.sort(key=lambda r: (-r["모델기여_합"], -r["총건수"], r["단어"]))
    return table


def _chi2_p(
    counter: Counter, base: Counter, present: int, total: int
) -> float | str:
    """등장 여부 × 라벨 2x3 표의 카이제곱 p값. 기대빈도가 낮으면 표시하지 않는다."""
    observed = [
        [counter[label] for label in LABELS],
        [base[label] - counter[label] for label in LABELS],
    ]
    if present == total or any(value < 0 for value in observed[1]):
        return ""
    try:
        result = chi2_contingency(observed)
    except ValueError:
        return ""
    return round(float(result.pvalue), 6)


def write_csv(table: Sequence[dict[str, Any]], path: Path) -> None:
    """엑셀이 한글을 바로 여는 utf-8-sig CSV로 저장한다."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(table[0]))
        writer.writeheader()
        writer.writerows(table)


def label_view(table: Sequence[dict[str, Any]], label: str) -> list[dict[str, Any]]:
    """한 라벨만 남긴 검토용 표. 그 라벨의 모델 기여가 큰 단어부터 정렬한다."""
    view = [
        {
            "단어": record["단어"],
            "어간근사": record["어간근사"],
            "총건수": record["총건수"],
            "문서수": record["문서수"],
            "건수": record[f"{label}_건수"],
            "비율": record[f"{label}_비율"],
            "lift": record[f"{label}_lift"],
            "chi2_p": record["chi2_p"],
            "모델기여": record[f"모델기여_{label}"],
            "최대lift_라벨": record["최대lift_라벨"],
        }
        for record in table
    ]
    view.sort(key=lambda r: (-r["모델기여"], -r["lift"], -r["총건수"], r["단어"]))
    return view


def write_xlsx(table: Sequence[dict[str, Any]], path: Path) -> None:
    """라벨별 시트로 나눈 엑셀 파일을 저장한다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    widths = {"단어": 22, "어간근사": 22, "최대lift_라벨": 16}
    formats = {"비율": "0.0%", "lift": "0.000", "모델기여": "0.0000", "chi2_p": "0.000000"}

    workbook = Workbook()
    workbook.remove(workbook.active)
    for label in LABELS:
        view = label_view(table, label)
        # 시트 이름에 '·'는 쓸 수 있으나 '/'·'*' 등은 엑셀이 거부하므로 미리 바꾼다.
        sheet = workbook.create_sheet(re.sub(r"[\\/*?\[\]:]", "_", label))
        headers = list(view[0])
        sheet.append(headers)
        for record in view:
            sheet.append([record[header] for header in headers])

        for index, header in enumerate(headers, start=1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = widths.get(header, 11)
            if header in formats:
                for cell in sheet[letter][1:]:
                    cell.number_format = formats[header]
            sheet[f"{letter}1"].font = Font(bold=True)
            sheet[f"{letter}1"].alignment = Alignment(horizontal="center")

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    version = os.getenv(DATASET_VERSION_ENV, DEFAULT_DATASET_KEY)
    default_dir = ROOT / "reports" / "current" / version
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--phrases",
        type=Path,
        default=default_dir / "model_explanation_phrases.csv",
        help="설명 보고서의 문구별 기여도 CSV. 없으면 기여도 열이 0으로 채워진다.",
    )
    parser.add_argument(
        "--output", type=Path, default=default_dir / "phrase_label_association.csv"
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=default_dir / "phrase_label_association.xlsx",
        help="라벨별 시트로 나눈 엑셀 파일. 세 라벨을 한 줄에 붙인 CSV도 함께 나온다.",
    )
    parser.add_argument("--min-count", type=int, default=3)
    args = parser.parse_args()

    rows, _ = load_label_dataset()
    excluded = evaluation_excluded_uids()
    rows = [row for row in rows if row["requirement_uid"] not in excluded]

    table = build_table(
        rows, load_model_contributions(args.phrases), min_count=args.min_count
    )
    print(f"단어 상관표 {len(table)}단어 / {len(rows)}건 기준")
    # 검토 중 파일이 엑셀에 열려 있으면 Windows가 쓰기를 막는다. 한쪽이 잠겨도
    # 다른 쪽은 갱신되도록 각각 따로 처리한다.
    for path, write in ((args.output, write_csv), (args.xlsx, write_xlsx)):
        try:
            write(table, path)
        except PermissionError:
            print(f"  건너뜀 (열려 있어 쓸 수 없음): {path}")
        else:
            print(f"  저장: {path}")


if __name__ == "__main__":
    main()
