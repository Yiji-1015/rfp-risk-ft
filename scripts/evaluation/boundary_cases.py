#!/usr/bin/env python3
"""`견적반영`과 `계약·질의검토`가 서로 뒤바뀐 건을 읽기용으로 모은다.

v4 OOF에서 단일 최고 모델(word+char Logistic)의 오답 294건 중 98건이 이 두 소수
클래스 사이의 상호 혼동이다. 두 라벨이 실제로 어디서 갈리는지는 점수가 아니라 라벨
생성 당시의 `reasoning`을 읽어야 알 수 있다. 이 스크립트는 그 읽을 거리를 만든다.

1위·2위 확률 차이가 작은 순으로 정렬한다. 모델이 가장 망설인 건이 경계가 가장
모호한 건이라는 가정이며, 그 가정 자체도 읽으면서 확인할 대상이다.
"""

from __future__ import annotations

import argparse
import csv
import os
from pathlib import Path
from typing import Any, Sequence

from scripts.evaluation.folds import evaluation_excluded_uids
from scripts.evaluation.phrase_label_association import write_csv
from scripts.labeling.label_dataset import (
    DATASET_VERSION_ENV,
    DEFAULT_DATASET_KEY,
    get_model_text,
    load_label_dataset,
)

ROOT = Path(__file__).resolve().parents[2]
BOUNDARY = frozenset({"견적반영", "계약·질의검토"})
DEFAULT_MODEL = "word_char_logistic"


def collect_boundary_cases(
    rows: Sequence[dict[str, Any]],
    oof_path: Path,
    *,
    model: str = DEFAULT_MODEL,
    labels: frozenset[str] = BOUNDARY,
) -> list[dict[str, Any]]:
    """두 라벨이 서로 뒤바뀐 건만 남겨 확률차 오름차순으로 돌려준다."""
    by_uid = {row["requirement_uid"]: row for row in rows}
    cases = []
    with oof_path.open(encoding="utf-8-sig", newline="") as handle:
        for saved in csv.DictReader(handle):
            gold, predicted = saved["gold"], saved[f"{model}_pred"]
            if gold == predicted or {gold, predicted} != labels:
                continue
            source = by_uid[saved["requirement_uid"]]
            probabilities = {
                label: float(saved[f"{model}_p_{label}"]) for label in labels
            }
            ordered = sorted(probabilities.values(), reverse=True)
            cases.append(
                {
                    "requirement_uid": saved["requirement_uid"],
                    "평가문서": saved["test_document"],
                    "정답": gold,
                    "예측": predicted,
                    "확률차": round(ordered[0] - ordered[1], 4),
                    **{
                        f"p_{label}": round(probabilities[label], 4)
                        for label in sorted(labels)
                    },
                    "요구사항명": source.get("requirement_name") or "",
                    "요구사항유형": source.get("requirement_type_normalized") or "",
                    "blockers": ", ".join(source.get("blockers") or []),
                    "cost_basis": source.get("cost_basis") or "",
                    "build_difficulty": source.get("build_difficulty") or "",
                    "domain_dependency": source.get("domain_dependency") or "",
                    "reasoning": source.get("reasoning") or "",
                    "본문": get_model_text(source),
                }
            )
    cases.sort(key=lambda case: (case["확률차"], case["requirement_uid"]))
    return cases


def write_xlsx(cases: Sequence[dict[str, Any]], path: Path) -> None:
    """긴 `reasoning`을 읽을 수 있게 줄바꿈과 열 너비를 준 엑셀로 저장한다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    widths = {
        "requirement_uid": 34,
        "평가문서": 24,
        "요구사항명": 34,
        "요구사항유형": 14,
        "blockers": 24,
        "cost_basis": 20,
        "reasoning": 90,
        "본문": 90,
    }
    wrapped = {"요구사항명", "blockers", "cost_basis", "reasoning", "본문"}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "견적↔계약 혼동"
    headers = list(cases[0])
    sheet.append(headers)
    for case in cases:
        sheet.append([case[header] for header in headers])

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(header, 12)
        sheet[f"{letter}1"].font = Font(bold=True)
        if header in wrapped:
            for cell in sheet[letter][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    version = os.getenv(DATASET_VERSION_ENV, DEFAULT_DATASET_KEY)
    default_dir = ROOT / "reports" / "current" / version
    parser = argparse.ArgumentParser()
    parser.add_argument("--oof", type=Path, default=default_dir / "model_candidate_oof.csv")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--output", type=Path, default=default_dir / "boundary_cases.csv")
    parser.add_argument("--xlsx", type=Path, default=default_dir / "boundary_cases.xlsx")
    args = parser.parse_args()

    rows, _ = load_label_dataset()
    excluded = evaluation_excluded_uids()
    rows = [row for row in rows if row["requirement_uid"] not in excluded]

    cases = collect_boundary_cases(rows, args.oof, model=args.model)
    print(f"견적반영 ↔ 계약·질의검토 상호 혼동 {len(cases)}건 ({args.model})")
    for path, write in ((args.output, write_csv), (args.xlsx, write_xlsx)):
        try:
            write(cases, path)
        except PermissionError:
            print(f"  건너뜀 (열려 있어 쓸 수 없음): {path}")
        else:
            print(f"  저장: {path}")


if __name__ == "__main__":
    main()
